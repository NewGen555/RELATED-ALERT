import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
import pandas as pd
import streamlit as st
from datetime import date
from PIL import Image
import gspread
from google.oauth2 import service_account

# =============================================================
# ตั้งค่าหน้าเว็บ Streamlit
# =============================================================
st.set_page_config(
    layout="wide",
    page_title="KFT Change Control System",
    page_icon="🔐"
)

TEMPLATE_FILE = "template_form.xlsx"
UPLOAD_DIR = "uploads"

os.makedirs(UPLOAD_DIR, exist_ok=True)

# =============================================================
# CONNECT GOOGLE SHEETS API
# =============================================================
@st.cache_resource
def get_gspread_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    
    json_path = os.path.join(os.path.dirname(__file__), "service_account.json")
    if os.path.exists(json_path):
        credentials = service_account.Credentials.from_service_account_file(
            json_path,
            scopes=scopes
        )
    else:
        creds_dict = dict(st.secrets["gcp_service_account"])
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
        credentials = service_account.Credentials.from_service_account_info(
            creds_dict,
            scopes=scopes
        )
        
    return gspread.authorize(credentials)

def get_worksheet():
    gc = get_gspread_client()
    spreadsheet_name = st.secrets.get("sheets", {}).get("spreadsheet_name", "change_control_db")
    sh = gc.open(spreadsheet_name)
    return sh.sheet1

# =============================================================
# CONFIGURATION: SMTP EMAIL SETTINGS & DEPARTMENT EMAILS
# =============================================================
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "pdd1development@gmail.com"
SENDER_PASSWORD = st.secrets.get("email", {}).get("sender_password", "")
APP_URL = "https://related-alert-erh2rywrtchautlthjlrwb.streamlit.app/"

DEPT_EMAILS = {
    "PDD": ["pdd_1@kftc.co.th", "saksiam@kftc.co.th", "manoc@kftc.co.th"],
    "QC": ["uchai@kftc.co.th", "sirirat@kftc.co.th", "pdd_1@kftc.co.th"],
    "PCD": ["pc-3@kftc.co.th", "pdd_1@kftc.co.th"],
    "PRD": ["suriya@kftc.co.th", "setthanan@kftc.co.th", "pd1center@kftc.co.th", "pdd_1@kftc.co.th"],
    "PRO": ["suriya@kftc.co.th", "setthanan@kftc.co.th", "pd1center@kftc.co.th", "pdd_1@kftc.co.th"],
    "PDD_MGR": ["manoch@kftc.co.th"],
    "QCD_MGR": ["maitree@kftc.co.th"],
    "PRD_MGR": ["suriya@kftc.co.th"],
    "PCD_MGR": ["umaporn@kftc.co.th"],
    "GM": ["mayuree@kftc.co.th"],
    "ALL": [
        "pdd_1@kftc.co.th", "saksiam@kftc.co.th", "manoc@kftc.co.th",
        "uchai@kftc.co.th", "sirirat@kftc.co.th", "pc-3@kftc.co.th",
        "suriya@kftc.co.th", "setthanan@kftc.co.th", "pd1center@kftc.co.th"
    ]
}

def send_email_notification(to_email, subject, body_content):
    recipient_list = to_email if isinstance(to_email, list) else [to_email]
    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To'] = ", ".join(recipient_list)
    msg['Subject'] = subject
    msg.attach(MIMEText(body_content, 'plain', 'utf-8'))

    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, recipient_list, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        st.error(f"❌ [DEBUG ERROR] ส่ง Email ไม่สำเร็จ สาเหตุเกิดจาก: {type(e).__name__} - {e}")
        return False

def send_all_completed_alert_email(doc_no, customer, part_name):
    to_email = DEPT_EMAILS.get("PDD_MGR", SENDER_EMAIL)
    subject = f"✅ [Wait Manager Approval] ใบงาน {doc_no} ปิดข้อ YES ครบถ้วนแล้ว (รอ PDD MGR อนุมัติ)"
    body = (
        f"เรียน ผู้จัดการ PDD (PDD MGR),\n\n"
        f"ใบงาน Change Control เลขที่ {doc_no} (Customer: {customer}, Part: {part_name}) "
        f"ได้รับการปิดข้อรายการที่ต้องแก้ไข (YES) พร้อมลงวันที่ปิดงานจริง (Actual Close) ครบถ้วนตาม Plan แล้ว\n\n"
        f"ระบบได้เปิดให้เข้าสู่ขั้นตอนอนุมัติแล้ว รบกวนผู้จัดการเข้าสู่ระบบเพื่อพิจารณาลงนามอนุมัติเอกสารเป็นลำดับแรกครับ\n\n"
        f"🔗 เข้าสู่ระบบได้ที่: {APP_URL}\n\n"
        f"ขอแสดงความนับถือ,\nระบบ KFT Change Control Automated System"
    )
    return send_email_notification(to_email, subject, body)

def send_approval_next_step_email(doc_no, customer, part_name, approver_title, next_approver_key, next_approver_title):
    to_email = DEPT_EMAILS.get(next_approver_key, SENDER_EMAIL)
    subject = f"🖊️ [Approval Step] ใบงาน {doc_no} รอการอนุมัติจาก {next_approver_title}"
    body = (
        f"เรียน {next_approver_title},\n\n"
        f"{approver_title} ได้ทำการลงนามอนุมัติเอกสารเลขที่ {doc_no} เรียบร้อยแล้ว\n"
        f"CUSTOMER: {customer}\nPART NAME: {part_name}\n\n"
        f"ลำดับถัดไป: รบกวนท่านเข้าสู่ระบบเพื่อพิจารณาอนุมัติเอกสารในระบบครับ\n\n"
        f"🔗 เข้าสู่ระบบได้ที่: {APP_URL}\n\n"
        f"ขอแสดงความนับถือ,\nระบบ KFT Change Control Automated System"
    )
    return send_email_notification(to_email, subject, body)

def send_final_approved_email(doc_no, customer, part_name, gm_name):
    to_email = DEPT_EMAILS.get("ALL", SENDER_EMAIL)
    subject = f"🎉 [FINAL APPROVED] เอกสาร Change Control {doc_no} ผ่านการอนุมัติเสร็จสมบูรณ์"
    body = (
        f"เรียน ทีมงานที่เกี่ยวข้องทุกท่าน,\n\n"
        f"เอกสารควบคุมการเปลี่ยนแปลง (Change Control) เลขที่ {doc_no}\n"
        f"CUSTOMER: {customer}\nPART NAME: {part_name}\n"
        f"ผู้อนุมัติขั้นสุดท้าย (GM): {gm_name}\n\n"
        f"บัดนี้ เอกสารดังกล่าวได้ผ่านการอนุมัติครบถ้วนตามลำดับขั้นตอนเรียบร้อยแล้วครับ\n\n"
        f"🔗 ตรวจสอบเอกสารได้ที่: {APP_URL}\n\n"
        f"ขอแสดงความนับถือ,\nระบบ KFT Change Control Automated System"
    )
    return send_email_notification(to_email, subject, body)

# =============================================================
# 🎨 CSS
# =============================================================
st.markdown("""
<style>
    .stApp { background: linear-gradient(135deg, #e3f2fd 0%, #ffffff 50%, #f0f8ff 100%); }
    h1, h2, h3 { color: #1565c0 !important; font-weight: 700 !important; }
    .login-container {
        background: rgba(255, 255, 255, 0.95); border-radius: 20px; padding: 40px;
        box-shadow: 0 8px 32px rgba(33, 150, 243, 0.15); border: 1px solid #bbdefb;
        max-width: 450px; margin: 80px auto; text-align: center;
    }
    .login-title { color: #1565c0; font-size: 28px; font-weight: 700; margin-bottom: 8px; }
    .login-subtitle { color: #64b5f6; font-size: 14px; margin-bottom: 30px; }
    .stButton > button {
        background: linear-gradient(135deg, #42a5f5 0%, #1976d2 100%) !important;
        color: white !important; border: none !important; border-radius: 12px !important;
        padding: 12px 24px !important; font-weight: 600 !important;
        box-shadow: 0 4px 15px rgba(25, 118, 210, 0.3) !important;
    }
    .status-card {
        padding: 15px; border-radius: 10px; background-color: #ffffff;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05); border-left: 5px solid #1565c0;
        margin-bottom: 10px;
    }
</style>
""", unsafe_allow_html=True)

# =============================================================
# 👤 USERS
# =============================================================
sec_passwords = st.secrets.get("passwords", {})

USERS = {
    "Thanawat": {"password": sec_passwords.get("Thanawat", ""), "dept": "PDD (Product Design)", "name": "ENGINEER PDD"},
    "qc_user": {"password": sec_passwords.get("qc_user", ""), "dept": "QC (Quality Control)", "name": "ENGINEER QC"},
    "pcd_user": {"password": sec_passwords.get("pcd_user", ""), "dept": "PCD (Production Control)", "name": "ENGINEER PCD"},
    "prd_user": {"password": sec_passwords.get("prd_user", ""), "dept": "PRO (Production / PD)", "name": "ENGINEER Production"},
    "mgr_pdd": {"password": sec_passwords.get("mgr_pdd", ""), "dept": "MGR - PDD (ผู้จัดการ PDD)", "name": "ผู้จัดการ PDD"},
    "mgr_qcd": {"password": sec_passwords.get("mgr_qcd", ""), "dept": "MGR - QCD (ผู้จัดการ QC)", "name": "ผู้จัดการ QC"},
    "mgr_pcd": {"password": sec_passwords.get("mgr_pcd", ""), "dept": "MGR - PCD (ผู้จัดการ PCD)", "name": "ผู้จัดการ PCD"},
    "mgr_prd": {"password": sec_passwords.get("mgr_prd", ""), "dept": "MGR - PD (ผู้จัดการ Production)", "name": "ผู้จัดการ PRD"},
    "gm_user": {"password": sec_passwords.get("gm_user", ""), "dept": "AGM / GM (ผู้บริหารอนุมัติขั้นสุดท้าย)", "name": "ผู้บริหาร GM"},
    "print_user": {"password": sec_passwords.get("print_user", ""), "dept": "Print Form", "name": "เจ้าหน้าที่พิมพ์เอกสาร"},
}

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.current_user = None
    st.session_state.current_dept = None
    st.session_state.user_name = None

def login(username, password):
    user = USERS.get(username)
    if user and user["password"] != "" and user["password"] == password:
        st.session_state.logged_in = True
        st.session_state.current_user = username
        st.session_state.current_dept = user["dept"]
        st.session_state.user_name = user["name"]
        return True
    return False

def logout():
    st.session_state.logged_in = False
    st.session_state.current_user = None
    st.session_state.current_dept = None
    st.session_state.user_name = None
    st.rerun()

# =============================================================
# 🔐 LOGIN UI
# =============================================================
if not st.session_state.logged_in:
    st.markdown("<style>[data-testid='stSidebar'] {display: none;}</style>", unsafe_allow_html=True)
    _, col2, _ = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
            <div class="login-container">
                <div style="font-size: 64px;">🔐</div>
                <div class="login-title">KFT Change Control</div>
                <div class="login-subtitle">ระบบควบคุมเอกสารการเปลี่ยนแปลง</div>
            </div>
        """, unsafe_allow_html=True)
        username = st.text_input("👤 Username", key="login_user")
        password = st.text_input("🔒 Password", type="password", key="login_pass")
        if st.button("🔓 เข้าสู่ระบบ", use_container_width=True, type="primary"):
            if login(username, password):
                st.success("✅ เข้าสู่ระบบสำเร็จ!")
                st.rerun()
            else:
                st.error("❌ Username หรือ Password ไม่ถูกต้อง")
    st.stop()

# =============================================================
# DATABASE OPERATIONS (GOOGLE SHEETS)
# =============================================================
ITEM_DEPT_MAPPING = {
    1: ("PDD", "MASTER DRAWING."), 2: ("PDD", "MATERIAL PART NO. LIST. , ACC DWG."),
    3: ("PDD", "PROCESS FLOW CHART."), 4: ("PDD", "OPERATION MANUAL."),
    5: ("PDD", "TEST RESULT."), 6: ("PDD", "FMEA"), 7: ("PDD", "TOOLING No"),
    8: ("QC", "CONTROL PLAN."), 9: ("QC", "INCOMING SHEET."),
    10: ("QC", "FINAL INSPECTION SHEET."), 11: ("QC", "W/I Out Going / TRAINING QC."),
    12: ("QC", "INSPECTION STD. + DATA CHECK."), 13: ("QC", "MSA"),
    14: ("QC", "PSW UP-DATE., PPAP APPROVAL."), 15: ("QC", "CHECKING FIXTURE."),
    16: ("PCD", "MATERIAL REQUIREMENT."), 17: ("PCD", "PACKING STANDARD."),
    18: ("PRO", "WORKING INSTRUCTION."), 19: ("PRO", "TRAINING PRODUCTION.")
}

def get_doc_value(doc_data, num, field_type):
    """ฟังก์ชันช่วยดึงค่าข้อมูลแต่ละข้อแบบยืดหยุ่น ป้องกันปัญหา Case-Sensitive และชื่อคอลัมน์ที่ไม่ตรงกัน"""
    if not doc_data:
        return ""
    keys_to_check = [
        f"DOC_{num}_{field_type.upper()}",
        f"DOC_{num}_{field_type.lower()}",
        f"doc_{num}_{field_type.lower()}",
        f"DOC{num}_{field_type.upper()}",
        f"DOC{num}_{field_type.lower()}"
    ]
    for key in keys_to_check:
        if key in doc_data and doc_data[key] is not None:
            return str(doc_data[key]).strip()
    return ""

def get_document_data(doc_no):
    try:
        ws = get_worksheet()
        records = ws.get_all_records()
        df = pd.DataFrame(records)
        
        if not df.empty:
            df.columns = [str(c).strip().upper() for c in df.columns]
            target_col = None
            for col in ['DOCUMENT_NO', 'DOCUMENT NO', 'DOC_NO']:
                if col in df.columns:
                    target_col = col
                    break
            
            if target_col:
                df[target_col] = df[target_col].astype(str).str.strip().str.upper()
                search_key = str(doc_no).strip().upper()
                matched = df[df[target_col] == search_key]
                if not matched.empty:
                    row_data = matched.iloc[0].to_dict()
                    return {str(k): ("" if pd.isna(v) else str(v).strip()) for k, v in row_data.items()}
    except Exception as e:
        st.error(f"เกิดข้อผิดพลาดในการอ่านข้อมูลจาก Google Sheets: {e}")
    return None

def get_all_documents():
    """ดึงข้อมูลแบบ Realtime สดจาก Google Sheets โดยไม่ผ่าน Cache"""
    try:
        ws = get_worksheet()
        records = ws.get_all_records()
        df = pd.DataFrame(records)
        if not df.empty:
            df.columns = [str(c).strip().upper() for c in df.columns]
            return df
    except Exception as e:
        st.error(f"เกิดข้อผิดพลาดในการดึงข้อมูล Realtime: {e}")
    return pd.DataFrame()

def save_to_excel(data_dict):
    try:
        ws = get_worksheet()
        headers = [str(h).strip().upper() for h in ws.row_values(1)]
        if not headers:
            st.error("❌ Google Sheet ยังไม่มี Header ในบรรทัดแรก")
            return False

        records = ws.get_all_records()
        df_old = pd.DataFrame(records)
        doc_no = str(data_dict.get('DOCUMENT_NO', '')).strip().upper()

        target_col = 'DOCUMENT_NO' if 'DOCUMENT_NO' in headers else headers[0]

        if not df_old.empty and target_col in df_old.columns and doc_no in df_old[target_col].astype(str).str.strip().str.upper().values:
            df_old[target_col] = df_old[target_col].astype(str).str.strip().str.upper()
            row_index = df_old[df_old[target_col] == doc_no].index[0] + 2
            cell_updates = []
            for key, value in data_dict.items():
                clean_key = str(key).strip().upper()
                if clean_key in headers and value is not None and value != "":
                    col_index = headers.index(clean_key) + 1
                    cell_updates.append(gspread.Cell(row=row_index, col=col_index, value=str(value)))
            if cell_updates:
                ws.update_cells(cell_updates)
        else:
            new_row = [str(data_dict.get(col, data_dict.get(col.upper(), ""))) for col in headers]
            ws.append_row(new_row)
        return True
    except Exception as e:
        st.error(f"❌ บันทึกข้อมูลลง Google Sheets ไม่สำเร็จ: {e}")
        return False

def check_yes_items_completed(doc_data):
    if not doc_data:
        return False, ["ไม่พบข้อมูลเอกสาร"]
    
    missing_list = []
    has_yes_item = False
    
    for num in range(1, 20):
        rev_val = get_doc_value(doc_data, num, "REVISE").upper()
        if rev_val == "YES":
            has_yes_item = True
            resp_val = get_doc_value(doc_data, num, "RESP")
            close_val = get_doc_value(doc_data, num, "CLOSE")
            dept, title = ITEM_DEPT_MAPPING.get(num, ("-", "-"))
            
            if not resp_val or resp_val == "-":
                missing_list.append(f"ข้อ {num} [{dept}]: ยังไม่ได้ลงชื่อผู้รับผิดชอบ ({title})")
            if not close_val or close_val == "-":
                missing_list.append(f"ข้อ {num} [{dept}]: ยังไม่ได้ลงวันที่ปิดเอกสารจริง ACTUAL CLOSE ({title})")
    
    if not has_yes_item:
        return True, []
        
    is_completed = (len(missing_list) == 0)
    return is_completed, missing_list

# =============================================================
# 🔍 ฟังก์ชันคำนวณตำแหน่งปัจจุบันของเอกสารแบบ Realtime
# =============================================================
def get_realtime_location(row):
    """วิเคราะห์สถานะและระบุแผนก/บุคคลที่ค้างเอกสารอยู่แบบละเอียด"""
    status = str(row.get('DOC_STATUS', '')).strip().upper()
    
    if status == "APPROVED" or row.get('APPR_GM'):
        return "🟢 อนุมัติเสร็จสมบูรณ์แล้ว", "อนุมัติครบถ้วน (GM Approved)", "SUCCESS"
    
    if not row.get('APPR_PDD_MGR'):
        return "🟡 รอการอนุมัติ", "อยู่ที่แผนก: PDD (รอ PDD Manager ลงนาม)", "MGR"
    elif not row.get('APPR_QCD_MGR'):
        return "🟡 รอการอนุมัติ", "อยู่ที่แผนก: QC (รอ QCD Manager ลงนาม)", "MGR"
    elif not row.get('APPR_PRD_MGR'):
        return "🟡 รอการอนุมัติ", "อยู่ที่แผนก: PRO/PD (รอ PRD Manager ลงนาม)", "MGR"
    elif not row.get('APPR_PCD_MGR'):
        return "🟡 รอการอนุมัติ", "อยู่ที่แผนก: PCD (รอ PCD Manager ลงนาม)", "MGR"
    elif not row.get('APPR_GM'):
        return "🟡 รอการอนุมัติ", "อยู่ที่ผู้บริหาร: AGM / GM (รอ GM ลงนามอนุมัติ)", "MGR"
        
    pending_depts = set()
    for num in range(1, 20):
        rev = get_doc_value(row, num, "REVISE").upper()
        if rev == "YES":
            close_val = get_doc_value(row, num, "CLOSE")
            resp_val = get_doc_value(row, num, "RESP")
            if not close_val or close_val == "-" or not resp_val or resp_val == "-":
                dept, _ = ITEM_DEPT_MAPPING.get(num, ("-", "-"))
                pending_depts.add(dept)
                
    if pending_depts:
        depts_str = ", ".join(sorted(list(pending_depts)))
        return "🔵 กำลังดำเนินการ", f"ติดอยู่ที่แผนก: {depts_str} (รอปิดข้อ YES & ลง Actual Close)", "ENGINEER"
        
    return "🔵 กำลังดำเนินการ", "อยู่ที่แผนก: PDD (รอยืนยันส่งต่อ Manager)", "ENGINEER"

def export_to_printed_form(doc_no):
    if not os.path.exists(TEMPLATE_FILE):
        return None, f"❌ ไม่พบไฟล์แบบฟอร์มต้นฉบับ '{TEMPLATE_FILE}' ในโฟลเดอร์โปรเจกต์"
        
    doc_data = get_document_data(doc_no)
    if not doc_data:
        return None, "❌ ไม่พบข้อมูลของเอกสารเลขที่นี้ในฐานข้อมูล"
        
    try:
        wb = openpyxl.load_workbook(TEMPLATE_FILE)
        ws = wb.active 
        
        def write_cell(coordinate, value):
            target_cell = ws[coordinate]
            for merged_range in list(ws.merged_cells.ranges):
                if target_cell.coordinate in merged_range:
                    ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
                    return
            target_cell.value = value
        
        write_cell("D3", doc_data.get("PART_NAME", ""))
        write_cell("D4", doc_data.get("PART_NO", ""))
        write_cell("F5", doc_data.get("MASTER_DWG_NO", ""))
        write_cell("P3", doc_data.get("MODEL", ""))
        write_cell("X3", doc_data.get("DATE", ""))
        write_cell("X1", doc_data.get("DOCUMENT_NO", ""))
        write_cell("H8", doc_data.get("REF_DOC_NO", ""))
        write_cell("X4", doc_data.get("ISSUE_BY", ""))
        
        write_cell("W7", doc_data.get("EFF_EVENT", ""))
        write_cell("W8", doc_data.get("EFF_PLAN", ""))
        write_cell("W9", doc_data.get("EFF_ACTUAL", ""))
        
        write_cell("D12", doc_data.get("SUBJECT_TEXT", ""))
        
        write_cell("I12", "X" if doc_data.get("ATTACH_DRAWING") == "YES" else "")
        write_cell("I13", "X" if doc_data.get("ATTACH_ECI") == "YES" else "")
        write_cell("I14", "X" if doc_data.get("ATTACH_MEETING") == "YES" else "")
        write_cell("I15", f"X ({doc_data.get('ATTACH_OTHERS_DETAIL', '')})" if doc_data.get("ATTACH_OTHERS") == "YES" else "")

        judgement_val = doc_data.get("JUDGEMENT", "")
        write_cell("S13", "X" if judgement_val == "FEASIBLE" else "")
        write_cell("S14", "X" if judgement_val == "IMPROBABILITY" else "")

        start_row = 19
        for i in range(1, 20):
            current_row = start_row + (i - 1)
            rev_val = get_doc_value(doc_data, i, "REVISE")
            if str(rev_val).upper() == "YES":
                write_cell(f"K{current_row}", "X")
                write_cell(f"M{current_row}", "")
            else:
                write_cell(f"K{current_row}", "")
                write_cell(f"M{current_row}", "X")
                
            write_cell(f"O{current_row}", get_doc_value(doc_data, i, "RESP"))
            write_cell(f"U{current_row}", get_doc_value(doc_data, i, "PLAN"))
            write_cell(f"Y{current_row}", get_doc_value(doc_data, i, "CLOSE"))
            
        write_cell("O41", doc_data.get("APPR_PDD_MGR", ""))
        write_cell("N44", doc_data.get("DATE_PDD_MGR", ""))
        write_cell("Q41", doc_data.get("APPR_QCD_MGR", ""))
        write_cell("Q44", doc_data.get("DATE_QCD_MGR", ""))
        write_cell("S41", doc_data.get("APPR_PCD_MGR", ""))
        write_cell("T44", doc_data.get("DATE_PCD_MGR", ""))
        write_cell("V41", doc_data.get("APPR_PRD_MGR", ""))
        write_cell("W44", doc_data.get("DATE_PRD_MGR", ""))
        write_cell("Y41", doc_data.get("APPR_GM", ""))
        write_cell("Z44", doc_data.get("DATE_GM", ""))
        
        safe_doc_no = doc_no.replace("/", "_").replace("\\", "_")
        output_filename = f"Change_Control_Sheet_{safe_doc_no}.xlsx"
        wb.save(output_filename)
        wb.close()
        return output_filename, None
    except Exception as e:
        return None, f"เกิดข้อผิดพลาดในการสร้างไฟล์ Excel: {str(e)}"

# =============================================================
# 📋 Sidebar Navigation
# =============================================================
with st.sidebar:
    st.markdown(f"👤 **{st.session_state.user_name}**\n\n🏢 **บทบาท:** {st.session_state.current_dept}")
    st.markdown("---")
    
    menu = st.radio("📌 เมนูใช้งาน", ["📊 Dashboard ติดตามสถานะ Realtime", "📝 บันทึก/อนุมัติ เอกสาร"])
    
    st.markdown("---")
    if st.button("🚪 ออกจากระบบ", use_container_width=True):
        logout()

selected_dept = st.session_state.current_dept

# =============================================================
# 📊 VIEW 1: DASHBOARD ติดตามสถานะแบบ REALTIME & ภาพรวม
# =============================================================
if menu == "📊 Dashboard ติดตามสถานะ Realtime":
    col_head1, col_head2 = st.columns([3, 1])
    with col_head1:
        st.title("📊 Realtime Tracking & Overview Dashboard")
        st.caption("ระบบติดตามตำแหน่งเอกสารแบบเรียลไทม์ และสรุปภาพรวมเอกสารทั้งหมดในระบบ")
    with col_head2:
        if st.button("🔄 รีเฟรชข้อมูลล่าสุด (Refresh Data)", use_container_width=True):
            st.rerun()

    df_all = get_all_documents()
    
    if df_all.empty:
        st.warning("⚠️ ยังไม่มีข้อมูลใบงานในระบบ")
    else:
        status_info = df_all.apply(get_realtime_location, axis=1)
        df_all['MAIN_STATUS'] = [s[0] for s in status_info]
        df_all['CURRENT_LOCATION'] = [s[1] for s in status_info]
        df_all['STAGE_TYPE'] = [s[2] for s in status_info]

        st.markdown("### 🔍 ค้นหาและติดตามตำแหน่งเอกสารแบบเจาะลึก (Realtime Search)")
        doc_search_input = st.text_input("กรอก DOCUMENT NO. ที่ต้องการติดตามตำแหน่ง (เช่น R001/26) :", placeholder="พิมพ์รหัสเอกสารที่นี่...").strip().upper()

        if doc_search_input:
            matched_df = df_all[df_all['DOCUMENT_NO'].astype(str).str.strip().str.upper() == doc_search_input]
            if not matched_df.empty:
                doc_row = matched_df.iloc[0]
                
                st.markdown(f"""
                <div class="status-card">
                    <h3 style="margin:0; color:#1565c0;">📄 เอกสารเลขที่: {doc_row['DOCUMENT_NO']}</h3>
                    <p style="margin:5px 0;"><b>Customer:</b> {doc_row.get('CUSTOMER_NAME', '-')} | <b>Part Name:</b> {doc_row.get('PART_NAME', '-')} | <b>Model:</b> {doc_row.get('MODEL', '-')}</p>
                    <h4 style="margin:10px 0 0 0; color:#d32f2f;">📍 สถานะปัจจุบัน: {doc_row['CURRENT_LOCATION']}</h4>
                </div>
                """, unsafe_allow_html=True)

                with st.expander("📌 **คลิกเพื่อดูรายละเอียดสถานะรายข้อ (Checklist 19 ข้อ & การเซ็น MGR)**", expanded=True):
                    col_t1, col_t2 = st.columns(2)
                    with col_t1:
                        st.markdown("##### 📝 **รายการที่เลือก YES และสถานะการปิดงาน:**")
                        has_yes = False
                        for num in range(1, 20):
                            rev = get_doc_value(doc_row, num, "REVISE").upper()
                            if rev == "YES":
                                has_yes = True
                                dept, title = ITEM_DEPT_MAPPING.get(num, ("-", "-"))
                                resp = get_doc_value(doc_row, num, "RESP")
                                close_dt = get_doc_value(doc_row, num, "CLOSE")
                                if resp and close_dt and close_dt != "-":
                                    st.success(f"✅ ข้อ {num} [{dept}]: {title} (ปิดงานเรียบร้อยโดย {resp} เมื่อ {close_dt})")
                                else:
                                    st.error(f"❌ ข้อ {num} [{dept}]: {title} (ยังไม่ปิดงาน - ค้างผู้รับผิดชอบ/Actual Close)")
                        if not has_yes:
                            st.info("ℹ️ ใบงานนี้ไม่มีหัวข้อที่เลือกแก้ไข (ไม่มีข้อ YES)")

                    with col_t2:
                        st.markdown("##### 🖊️ **สถานะการลงนามอนุมัติ (Manager Approval):**")
                        st.write(f"1. **PDD MGR:** {doc_row.get('APPR_PDD_MGR') if doc_row.get('APPR_PDD_MGR') else '⏳ รอการลงนาม'}")
                        st.write(f"2. **QCD MGR:** {doc_row.get('APPR_QCD_MGR') if doc_row.get('APPR_QCD_MGR') else '⏳ รอการลงนาม'}")
                        st.write(f"3. **PRD MGR:** {doc_row.get('APPR_PRD_MGR') if doc_row.get('APPR_PRD_MGR') else '⏳ รอการลงนาม'}")
                        st.write(f"4. **PCD MGR:** {doc_row.get('APPR_PCD_MGR') if doc_row.get('APPR_PCD_MGR') else '⏳ รอการลงนาม'}")
                        st.write(f"5. **AGM / GM:** {doc_row.get('APPR_GM') if doc_row.get('APPR_GM') else '⏳ รอการลงนาม'}")
            else:
                st.error(f"❌ ไม่พบเอกสารเลขที่ '{doc_search_input}' ในฐานข้อมูล")
        
        st.markdown("---")

        st.markdown("### 📈 ภาพรวมเอกสารทั้งหมดในระบบ (System Overview)")
        
        total_docs = len(df_all)
        approved_docs = len(df_all[df_all['MAIN_STATUS'].str.contains("Approved")])
        pending_mgr = len(df_all[df_all['MAIN_STATUS'].str.contains("รอการอนุมัติ")])
        in_progress = total_docs - approved_docs - pending_mgr

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("📄 ใบงานทั้งหมดในระบบ", f"{total_docs} รายการ")
        m2.metric("🔵 รอแผนกปิดข้อ YES", f"{in_progress} รายการ")
        m3.metric("🟡 รอผู้จัดการอนุมัติ (Wait MGR)", f"{pending_mgr} รายการ")
        m4.metric("🟢 อนุมัติเสร็จสมบูรณ์", f"{approved_docs} รายการ")

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("##### 📋 **ตารางติดตามตำแหน่งเอกสารแบบเรียลไทม์ (Realtime Document List)**")
        
        c_filter1, c_filter2 = st.columns(2)
        with c_filter1:
            filter_status = st.selectbox("🎯 กรองตามสถานะหลัก:", ["ทั้งหมด", "🔵 กำลังดำเนินการ", "🟡 รอการอนุมัติ", "🟢 อนุมัติเสร็จสมบูรณ์"])
        with c_filter2:
            filter_text = st.text_input("🔍 กรองตาม Customer / Part Name / Issue By:", key="table_search")

        display_df = df_all.copy()
        
        if filter_status != "ทั้งหมด":
            display_df = display_df[display_df['MAIN_STATUS'].str.contains(filter_status[:2])]
            
        if filter_text:
            ft = filter_text.strip().upper()
            display_df = display_df[
                display_df['CUSTOMER_NAME'].astype(str).str.upper().str.contains(ft) |
                display_df['PART_NAME'].astype(str).str.upper().str.contains(ft) |
                display_df['ISSUE_BY'].astype(str).str.upper().str.contains(ft)
            ]

        show_cols = ['DOCUMENT_NO', 'CUSTOMER_NAME', 'PART_NAME', 'MODEL', 'ISSUE_BY', 'MAIN_STATUS', 'CURRENT_LOCATION']
        valid_cols = [c for c in show_cols if c in display_df.columns]

        st.dataframe(
            display_df[valid_cols].rename(columns={
                'DOCUMENT_NO': 'เลขที่เอกสาร',
                'CUSTOMER_NAME': 'ลูกค้า',
                'PART_NAME': 'ชื่อพาร์ท',
                'MODEL': 'โมเดล',
                'ISSUE_BY': 'ผู้จัดทำ',
                'MAIN_STATUS': 'สถานะหลัก',
                'CURRENT_LOCATION': '📍 ตำแหน่งปัจจุบันของเอกสาร (ผู้รับผิดชอบ)'
            }),
            use_container_width=True,
            hide_index=True
        )

# =============================================================
# 📝 VIEW 2: หน้าบันทึก / อนุมัติ เอกสาร
# =============================================================
else:
    st.title("KFT - RELATED DOCUMENT CHANGE CONTROL SYSTEM")

    if "Print Form" in selected_dept:
        st.subheader("🖨️ ระบบดึงและพิมพ์ฟอร์มเอกสารควบคุมอัตโนมัติ (Excel Format บริษัท)")
        print_doc_no = st.text_input("กรอก DOCUMENT NO. ที่ต้องการแปลงข้อมูลออกฟอร์ม (เช่น R001/26) :", key="print_doc_input").strip().upper()
        
        if print_doc_no:
            doc_data = get_document_data(print_doc_no)
            if doc_data:
                st.success(f"พบข้อมูลของใบงานหมายเลข {print_doc_no} ในระบบ")
                st.info(f"📋 สรุปงาน -> ลูกค้า: {doc_data.get('CUSTOMER_NAME', '-')} | พาร์ท: {doc_data.get('PART_NAME', '-')}")
                
                output_file, error_msg = export_to_printed_form(print_doc_no)
                if error_msg:
                    st.error(error_msg)
                elif output_file and os.path.exists(output_file):
                    with open(output_file, "rb") as f:
                        st.download_button(
                            label="📥 คลิกตรงนี้เพื่อดาวน์โหลดแบบฟอร์มจริงสำเร็จรูป (.xlsx)",
                            data=f,
                            file_name=os.path.basename(output_file),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True
                        )
                    st.balloons()
            else:
                st.error("❌ ไม่พบข้อมูลรหัสใบงานนี้ในฐานข้อมูลระบบ")

    elif "MGR" in selected_dept or "GM" in selected_dept:
        st.subheader("🔒 ระบบพิจารณาลงนามดิจิทัลและอนุมัติเอกสารควบคุม (Manager Approval Loop)")
        approve_doc_no = st.text_input("ระบุ DOCUMENT NO. ที่ต้องการพิจารณาอนุมัติ :", key="appr_doc_input").strip().upper()
        
        if approve_doc_no:
            doc_data = get_document_data(approve_doc_no)
            if doc_data:
                cust = doc_data.get('CUSTOMER_NAME', '-')
                part = doc_data.get('PART_NAME', '-')
                doc_status = doc_data.get('DOC_STATUS', 'PENDING')

                st.info(f"📋 รายละเอียดใบงาน -> ลูกค้า: {cust} | ชื่อพาร์ท: {part} | สถานะเอกสาร: {doc_status}")
                
                is_completed, missing_items = check_yes_items_completed(doc_data)

                if not is_completed:
                    st.error("⛔ [ระบบล็อกการอนุมัติ] ใบงานนี้ยังไม่สามารถเข้าสู่ลูปการอนุมัติระดับ Manager ได้")
                    st.warning("สาเหตุ: พนักงานแผนกที่เกี่ยวข้องยังไม่ได้กรอกข้อมูลให้ครบถ้วน ดังนี้:")
                    for missing in missing_items:
                        st.write(f"- {missing}")
                else:
                    st.success("✅ รายการที่ต้องแก้ไข (YES) ปิดงานครบถ้วนแล้ว พร้อมสำหรับการอนุมัติ")
                    
                    pdd_appr = doc_data.get('APPR_PDD_MGR', '')
                    qcd_appr = doc_data.get('APPR_QCD_MGR', '')
                    prd_appr = doc_data.get('APPR_PRD_MGR', '')
                    pcd_appr = doc_data.get('APPR_PCD_MGR', '')
                    gm_appr = doc_data.get('APPR_GM', '')

                    col_s1, col_s2, col_s3, col_s4, col_s5 = st.columns(5)
                    col_s1.metric("1. PDD MGR", pdd_appr if pdd_appr else "⏳ รอลงนาม")
                    col_s2.metric("2. QCD MGR", qcd_appr if qcd_appr else "⏳ รอลงนาม")
                    col_s3.metric("3. PRD MGR", prd_appr if prd_appr else "⏳ รอลงนาม")
                    col_s4.metric("4. PCD MGR", pcd_appr if pcd_appr else "⏳ รอลงนาม")
                    col_s5.metric("5. AGM/GM", gm_appr if gm_appr else "⏳ รอลงนาม")

                    can_approve = False
                    approver_role_key = ""
                    next_role_key = ""
                    next_role_title = ""

                    if "MGR - PDD" in selected_dept:
                        can_approve = True
                        approver_role_key = "APPR_PDD_MGR"
                        date_role_key = "DATE_PDD_MGR"
                        next_role_key = "QCD_MGR"
                        next_role_title = "ผู้จัดการ QCD (Quality Control)"
                    elif "MGR - QCD" in selected_dept:
                        if pdd_appr:
                            can_approve = True
                            approver_role_key = "APPR_QCD_MGR"
                            date_role_key = "DATE_QCD_MGR"
                            next_role_key = "PRD_MGR"
                            next_role_title = "ผู้จัดการ PRD/PRO (Production)"
                        else:
                            st.warning("⚠️ รอ PDD MGR ลงนามอนุมัติก่อนหน้าตามลำดับลำดับขั้น")
                    elif "MGR - PD" in selected_dept:
                        if qcd_appr:
                            can_approve = True
                            approver_role_key = "APPR_PRD_MGR"
                            date_role_key = "DATE_PRD_MGR"
                            next_role_key = "PCD_MGR"
                            next_role_title = "ผู้จัดการ PCD (Production Control)"
                        else:
                            st.warning("⚠️ รอ QCD MGR ลงนามอนุมัติก่อนหน้าตามลำดับขั้น")
                    elif "MGR - PCD" in selected_dept:
                        if prd_appr:
                            can_approve = True
                            approver_role_key = "APPR_PCD_MGR"
                            date_role_key = "DATE_PCD_MGR"
                            next_role_key = "GM"
                            next_role_title = "ผู้บริหาร AGM / GM"
                        else:
                            st.warning("⚠️ รอ PRD MGR ลงนามอนุมัติก่อนหน้าตามลำดับขั้น")
                    elif "AGM / GM" in selected_dept:
                        if pcd_appr:
                            can_approve = True
                            approver_role_key = "APPR_GM"
                            date_role_key = "DATE_GM"
                        else:
                            st.warning("⚠️ รอ PCD MGR ลงนามอนุมัติก่อนหน้าตามลำดับขั้น")

                    if can_approve:
                        st.markdown("---")
                        st.subheader(f"🖊️ ส่วนลงนามอนุมัติ: {st.session_state.current_dept}")
                        approver_name = st.text_input("กรอกชื่อ-นามสกุล ผู้ลงนามอนุมัติ:", value=st.session_state.user_name)
                        appr_date = st.date_input("วันที่ลงนามอนุมัติ:", value=date.today())

                        if st.button("✅ ยืนยันการลงนามอนุมัติเอกสาร", type="primary", use_container_width=True):
                            update_dict = {
                                "DOCUMENT_NO": approve_doc_no,
                                approver_role_key: approver_name,
                                date_role_key: str(appr_date)
                            }
                            
                            if approver_role_key == "APPR_GM":
                                update_dict["DOC_STATUS"] = "APPROVED"
                            
                            if save_to_excel(update_dict):
                                st.success("🎉 ลงนามอนุมัติเอกสารเรียบร้อยแล้ว!")
                                
                                if next_role_key:
                                    send_approval_next_step_email(
                                        approve_doc_no, cust, part, 
                                        st.session_state.current_dept, 
                                        next_role_key, next_role_title
                                    )
                                elif approver_role_key == "APPR_GM":
                                    send_final_approved_email(approve_doc_no, cust, part, approver_name)
                                    
                                st.rerun()
            else:
                st.error("❌ ไม่พบข้อมูลรหัสเอกสารนี้ในระบบ")

    else:
        # =============================================================
        # ส่วนงานบันทึก/แก้ไขข้อมูลของ Engineer (PDD, QC, PCD, PRO)
        # =============================================================
        st.subheader(f"📝 แบบฟอร์มกรอกข้อมูลการเปลี่ยนแปลง ({selected_dept})")
        
        doc_no = st.text_input("📌 DOCUMENT NO. (เช่น R001/26):", key="main_doc_no").strip().upper()
        doc_data = get_document_data(doc_no) if doc_no else {}

        if doc_no and doc_data:
            st.info(f"ℹ️ พบข้อมูลเดิมของใบงาน {doc_no} ระบบจะแสดงข้อมูลล่าสุดให้อัตโนมัติ")
        elif doc_no:
            st.warning(f"🆕 ไม่พบใบงาน {doc_no} ในระบบ (จะเป็นการสร้างใบงานใหม่)")

        # Section 1: ข้อมูลทั่วไป (PDD กรอก)
        if "PDD" in selected_dept:
            st.markdown("#### 1. ข้อมูลทั่วไปของเอกสาร (General Information)")
            c1, c2, c3 = st.columns(3)
            with c1:
                customer_name = st.text_input("CUSTOMER NAME", value=doc_data.get("CUSTOMER_NAME", ""), key=f"cust_{doc_no}")
                part_name = st.text_input("PART NAME", value=doc_data.get("PART_NAME", ""), key=f"partname_{doc_no}")
                part_no = st.text_input("PART NO.", value=doc_data.get("PART_NO", ""), key=f"partno_{doc_no}")
            with c2:
                model = st.text_input("MODEL", value=doc_data.get("MODEL", ""), key=f"model_{doc_no}")
                master_dwg_no = st.text_input("MASTER DWG NO.", value=doc_data.get("MASTER_DWG_NO", ""), key=f"dwg_{doc_no}")
                ref_doc_no = st.text_input("REF. DOC NO.", value=doc_data.get("REF_DOC_NO", ""), key=f"refdoc_{doc_no}")
            with c3:
                doc_date = st.text_input("DATE (yyyy-mm-dd)", value=doc_data.get("DATE", str(date.today())), key=f"date_{doc_no}")
                issue_by = st.text_input("ISSUE BY", value=doc_data.get("ISSUE_BY", st.session_state.user_name), key=f"issue_{doc_no}")
                subject_text = st.text_area("SUBJECT / รายละเอียดการเปลี่ยนแปลง", value=doc_data.get("SUBJECT_TEXT", ""), key=f"subj_{doc_no}")

            st.markdown("#### 2. กำหนดการ Effective Date & เอกสารแนบ")
            ce1, ce2, ce3 = st.columns(3)
            with ce1:
                eff_event = st.text_input("EFFECTIVE EVENT", value=doc_data.get("EFF_EVENT", ""), key=f"eff_event_{doc_no}")
            with ce2:
                eff_plan = st.text_input("EFFECTIVE PLAN DATE", value=doc_data.get("EFF_PLAN", ""), key=f"eff_plan_{doc_no}")
            with ce3:
                eff_actual = st.text_input("EFFECTIVE ACTUAL DATE", value=doc_data.get("EFF_ACTUAL", ""), key=f"eff_act_{doc_no}")

            ca1, ca2, ca3, ca4 = st.columns(4)
            with ca1:
                attach_dwg = st.checkbox("DRAWING", value=(doc_data.get("ATTACH_DRAWING") == "YES"), key=f"at_dwg_{doc_no}")
            with ca2:
                attach_eci = st.checkbox("ECI", value=(doc_data.get("ATTACH_ECI") == "YES"), key=f"at_eci_{doc_no}")
            with ca3:
                attach_mtg = st.checkbox("MEETING MEMO", value=(doc_data.get("ATTACH_MEETING") == "YES"), key=f"at_mtg_{doc_no}")
            with ca4:
                attach_oth = st.checkbox("OTHERS", value=(doc_data.get("ATTACH_OTHERS") == "YES"), key=f"at_oth_{doc_no}")
                attach_oth_detail = st.text_input("ระบุ OTHERS", value=doc_data.get("ATTACH_OTHERS_DETAIL", ""), key=f"at_oth_dt_{doc_no}")

            judgement_val = st.radio("JUDGEMENT RESULT", ["FEASIBLE", "IMPROBABILITY"], 
                                     index=0 if doc_data.get("JUDGEMENT") != "IMPROBABILITY" else 1, key=f"judge_{doc_no}")

        # Section 2: รายการ Checklist 19 ข้อ แบ่งตามแผนก
        st.markdown("---")
        st.subheader("📋 รายการเอกสารที่ต้องแก้ไข ลงนาม และปิดเอกสารตาม Plan (Checklist 19 ข้อ)")

        # ฟังก์ชัน Helper สร้าง UI สำหรับแต่ละข้อ (ปรับปรุง key ให้ผูกกับ doc_no เพื่ออัปเดตค่าตามจริง)
        def render_checklist_item(num, title, doc_data, current_doc_no):
            rev_raw = get_doc_value(doc_data, num, "REVISE").upper()
            resp_val = get_doc_value(doc_data, num, "RESP")
            plan_val = get_doc_value(doc_data, num, "PLAN")
            close_val = get_doc_value(doc_data, num, "CLOSE")

            radio_index = 1 if rev_raw == "YES" else 0
            doc_key_prefix = current_doc_no if current_doc_no else "new"

            st.markdown(f"**ข้อ {num}. {title}**")
            c1, c2, c3, c4 = st.columns([1.5, 2.5, 2, 2])
            with c1:
                rev_input = st.radio(f"แก้ไข? ({num})", options=["NO", "YES"], index=radio_index, key=f"rev_{doc_key_prefix}_{num}")
            with c2:
                resp_input = st.text_input(f"ผู้รับผิดชอบ ({num})", value=resp_val, key=f"resp_{doc_key_prefix}_{num}")
            with c3:
                plan_input = st.text_input(f"กำหนดเสร็จ PLAN ({num})", value=plan_val, key=f"plan_{doc_key_prefix}_{num}")
            with c4:
                close_input = st.text_input(f"วันที่ปิดจริง ACTUAL ({num})", value=close_val, key=f"close_{doc_key_prefix}_{num}")
            st.markdown("---")
            return rev_input, resp_input, plan_input, close_input

        checklist_results = {}

        # ------------------- PDD SECTION (ข้อ 1 - 7) -------------------
        if "PDD" in selected_dept:
            st.markdown("### 🔹 ส่วนงาน PDD (ข้อ 1 - 7)")
            pdd_items = [
                (1, "MASTER DRAWING."), (2, "MATERIAL PART NO. LIST. , ACC DWG."),
                (3, "PROCESS FLOW CHART."), (4, "OPERATION MANUAL."),
                (5, "TEST RESULT."), (6, "FMEA"), (7, "TOOLING No")
            ]
            for num, title in pdd_items:
                checklist_results[num] = render_checklist_item(num, title, doc_data, doc_no)

        # ------------------- QC SECTION (ข้อ 8 - 15) -------------------
        elif "QC" in selected_dept:
            st.markdown("### 🔹 ส่วนงาน QC (ข้อ 8 - 15)")
            qc_items = [
                (8, "CONTROL PLAN."), (9, "INCOMING SHEET."),
                (10, "FINAL INSPECTION SHEET."), (11, "W/I Out Going / TRAINING QC."),
                (12, "INSPECTION STD. + DATA CHECK."), (13, "MSA"),
                (14, "PSW UP-DATE., PPAP APPROVAL."), (15, "CHECKING FIXTURE.")
            ]
            for num, title in qc_items:
                checklist_results[num] = render_checklist_item(num, title, doc_data, doc_no)

        # ------------------- PCD SECTION (ข้อ 16 - 17) -------------------
        elif "PCD" in selected_dept:
            st.markdown("### 🔹 ส่วนงาน PCD (ข้อ 16 - 17)")
            pcd_items = [
                (16, "MATERIAL REQUIREMENT."), (17, "PACKING STANDARD.")
            ]
            for num, title in pcd_items:
                checklist_results[num] = render_checklist_item(num, title, doc_data, doc_no)

        # ------------------- PRO SECTION (ข้อ 18 - 19) -------------------
        elif "PRO" in selected_dept:
            st.markdown("### 🔹 ส่วนงาน Production / PRO (ข้อ 18 - 19)")
            pro_items = [
                (18, "WORKING INSTRUCTION."), (19, "TRAINING PRODUCTION.")
            ]
            for num, title in pro_items:
                checklist_results[num] = render_checklist_item(num, title, doc_data, doc_no)

        # =============================================================
        # ปุ่มบันทึกข้อมูล (Save Operations)
        # =============================================================
        if st.button("💾 บันทึกข้อมูลลงระบบ", type="primary", use_container_width=True):
            if not doc_no:
                st.error("❌ กรุณากรอก DOCUMENT NO. ก่อนทำการบันทึกข้อมูล")
            else:
                save_payload = {"DOCUMENT_NO": doc_no}

                # รวมข้อมูล General หากเป็น PDD
                if "PDD" in selected_dept:
                    save_payload.update({
                        "CUSTOMER_NAME": customer_name,
                        "PART_NAME": part_name,
                        "PART_NO": part_no,
                        "MODEL": model,
                        "MASTER_DWG_NO": master_dwg_no,
                        "REF_DOC_NO": ref_doc_no,
                        "DATE": str(doc_date),
                        "ISSUE_BY": issue_by,
                        "SUBJECT_TEXT": subject_text,
                        "EFF_EVENT": eff_event,
                        "EFF_PLAN": eff_plan,
                        "EFF_ACTUAL": eff_actual,
                        "ATTACH_DRAWING": "YES" if attach_dwg else "NO",
                        "ATTACH_ECI": "YES" if attach_eci else "NO",
                        "ATTACH_MEETING": "YES" if attach_mtg else "NO",
                        "ATTACH_OTHERS": "YES" if attach_oth else "NO",
                        "ATTACH_OTHERS_DETAIL": attach_oth_detail if attach_oth else "",
                        "JUDGEMENT": judgement_val
                    })

                # รวมข้อมูลรายการ Checklist ที่แก้ไข
                for num, (r_val, resp_val, plan_val, close_val) in checklist_results.items():
                    save_payload[f"DOC_{num}_REVISE"] = r_val
                    save_payload[f"DOC_{num}_RESP"] = resp_val
                    save_payload[f"DOC_{num}_PLAN"] = plan_val
                    save_payload[f"DOC_{num}_CLOSE"] = close_val

                # ดำเนินการบันทึก
                if save_to_excel(save_payload):
                    st.success(f"✅ บันทึกข้อมูลใบงาน {doc_no} สำเร็จเรียบร้อยแล้ว!")
                    
                    # ตรวจสอบสถานะว่าทุกข้อ YES ปิดครบหมดแล้วหรือยัง เพื่อส่ง Email แจ้ง PDD MGR
                    updated_doc_data = get_document_data(doc_no)
                    if updated_doc_data:
                        is_all_complete, _ = check_yes_items_completed(updated_doc_data)
                        if is_all_complete and not updated_doc_data.get('APPR_PDD_MGR'):
                            send_all_completed_alert_email(
                                doc_no, 
                                updated_doc_data.get('CUSTOMER_NAME', '-'), 
                                updated_doc_data.get('PART_NAME', '-')
                            )
                            st.info("📧 ส่งอีเมลแจ้งเตือนไปยัง PDD Manager เพื่อรอพิจารณาอนุมัติเรียบร้อยแล้ว")

                    st.rerun()
